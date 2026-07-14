"""Export a table (optionally a row subset) to Parquet, CSV, Excel, JSON, or Feather.

These operate on an in-memory/mmap :class:`pyarrow.Table`, which is what the interactive UI
holds. The optional ``rows`` argument carries the filtered-view or user-selected row indices,
so the same call powers "export all", "export filtered view", and "export selected rows".
"""

from __future__ import annotations

import json
import logging
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any, Literal

import pyarrow as pa
import pyarrow.csv as pa_csv
import pyarrow.parquet as pq

from .model import EXCEL_MAX_ROWS
from .tables import TableStore

logger = logging.getLogger(__name__)

ExportFormat = Literal["parquet", "csv", "xlsx", "json", "feather"]
_FORMATS: frozenset[str] = frozenset({"parquet", "csv", "xlsx", "json", "feather"})

# Map a format to its conventional file extension.
FORMAT_SUFFIX: dict[str, str] = {
    "parquet": ".parquet",
    "csv": ".csv",
    "xlsx": ".xlsx",
    "json": ".json",
    "feather": ".arrow",
}


def _subset(table: pa.Table, rows: Sequence[int] | None) -> pa.Table:
    """Return ``table`` restricted to ``rows`` (or the whole table when ``rows`` is None)."""
    if rows is None:
        return table
    return table.take(pa.array(rows, type=pa.int64()))


def _fill_sheet(worksheet: object, table: pa.Table) -> int:
    """Append ``table`` (header + rows) to an openpyxl write-only sheet; return rows written.

    Streams row-wise per batch so memory stays bounded, and stops at Excel's row cap.
    """
    usable = EXCEL_MAX_ROWS - 1
    columns = table.column_names
    worksheet.append(columns)  # type: ignore[attr-defined]  # openpyxl WriteOnlyWorksheet

    written = 0
    for batch in table.to_batches():
        for record in batch.to_pylist():  # row-wise dicts; keeps memory bounded per batch
            if written >= usable:
                logger.warning(
                    "xlsx sheet truncated to Excel's row limit (%d); use parquet/csv/feather.",
                    usable,
                )
                return written
            worksheet.append([record[name] for name in columns])  # type: ignore[attr-defined]
            written += 1
    return written


def _write_xlsx(table: pa.Table, path: Path, sheet_name: str) -> int:
    """Write ``table`` to a single-sheet xlsx via streaming openpyxl; return rows written."""
    from openpyxl import Workbook  # noqa: PLC0415 - lazy; xlsx-only cost

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name[:31] or "data")  # Excel sheet-name limit is 31 chars
    written = _fill_sheet(ws, table)
    wb.save(str(path))
    return written


def export(
    table: pa.Table,
    path: Path,
    fmt: ExportFormat,
    *,
    rows: Sequence[int] | None = None,
    sheet_name: str = "data",
) -> int:
    """Export ``table`` (optionally only ``rows``) to ``path`` in ``fmt``.

    Args:
        table: The source table.
        path: Destination file path.
        fmt: One of parquet, csv, xlsx, json, feather.
        rows: Optional row indices to export (filtered view or selection); None exports all.
        sheet_name: Sheet title for the xlsx format.

    Returns:
        The number of rows written.

    Raises:
        ValueError: If ``fmt`` is not a supported format.
    """
    if fmt not in _FORMATS:
        raise ValueError(f"unsupported format {fmt!r} (use: {', '.join(sorted(_FORMATS))})")
    subset = _subset(table, rows)
    path.parent.mkdir(parents=True, exist_ok=True)

    match fmt:
        case "parquet":
            pq.write_table(subset, str(path))  # pyright: ignore[reportUnknownMemberType]  # stub embeds Unknown
        case "feather":
            # Arrow IPC file format (Feather v2), written via the non-deprecated ipc API.
            with pa.OSFile(str(path), "wb") as sink, pa.ipc.new_file(sink, subset.schema) as writer:
                writer.write_table(subset)
        case "csv":
            pa_csv.write_csv(_stringify_list_columns(subset), str(path))
        case "json":
            path.write_text(json.dumps(subset.to_pylist(), ensure_ascii=False), encoding="utf-8")
        case "xlsx":
            return _write_xlsx(_stringify_list_columns(subset), path, sheet_name)
    return subset.num_rows


def _stringify_list_columns(table: pa.Table) -> pa.Table:
    """Join any list-typed column into a ``"; "``-delimited string.

    CSV and Excel cells cannot hold arrays, but the CPC steps produce list columns
    (``cpc_codes``, ``cpc_subclasses``, ``shared_codes``). Parquet/Feather/JSON keep them as lists;
    the text formats flatten each list to a readable string (``None`` → empty).
    """
    out = table
    for name in table.column_names:
        column = out.column(name)
        if not (pa.types.is_list(column.type) or pa.types.is_large_list(column.type)):
            continue
        items: list[Any] = column.to_pylist()
        joined = ["; ".join(str(v) for v in item) if item else "" for item in items]
        out = out.set_column(
            out.schema.get_field_index(name), name, pa.array(joined, type=pa.string())
        )
    return out


def write_workbook(path: Path, sheets: Mapping[str, pa.Table]) -> dict[str, int]:
    """Write a multi-sheet xlsx workbook (one sheet per entry, streamed).

    Sheet titles are truncated to Excel's 31-character limit. Returns rows written per sheet.
    """
    from openpyxl import Workbook  # noqa: PLC0415 - lazy; xlsx-only cost

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=True)
    counts = {
        name: _fill_sheet(wb.create_sheet(title=name[:31]), table) for name, table in sheets.items()
    }
    wb.save(str(path))
    return counts


def export_store(store: TableStore, out_dir: Path, fmt: ExportFormat) -> dict[str, int]:
    """Export every table in ``store`` into ``out_dir`` in ``fmt``.

    For ``xlsx`` this writes a single multi-sheet workbook (one sheet per table); for the other
    formats it writes one file per table (``<table>.<ext>``).

    Args:
        store: The dataset to export.
        out_dir: Directory to write into (created if missing).
        fmt: One of parquet, csv, xlsx, json, feather.

    Returns:
        Mapping of table name to rows written.

    Raises:
        ValueError: If ``fmt`` is not a supported format.
    """
    if fmt not in _FORMATS:
        raise ValueError(f"unsupported format {fmt!r} (use: {', '.join(sorted(_FORMATS))})")
    out_dir.mkdir(parents=True, exist_ok=True)

    if fmt == "xlsx":
        return write_workbook(
            out_dir / "assignments.xlsx", {name: store.table(name) for name in store.names}
        )

    return {
        name: export(store.table(name), out_dir / f"{name}{FORMAT_SUFFIX[fmt]}", fmt)
        for name in store.names
    }
