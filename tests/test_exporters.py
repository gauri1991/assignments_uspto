"""Tests for multi-format export with All / subset scope (uspto_assignments.exporters)."""

from __future__ import annotations

import json
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from openpyxl import load_workbook

from uspto_assignments import TableStore, exporters, parse_to_store

FIXTURE = Path(__file__).parent / "fixtures" / "sample_assignment.xml"


def _table() -> pa.Table:
    return pa.table(
        {
            "reel_no": ["012345", "054321", "000777"],
            "name": ["ACME", "BETA", "GAMMA"],
        }
    )


def test_export_parquet_roundtrip_preserves_leading_zeros(tmp_path: Path) -> None:
    n = exporters.export(_table(), tmp_path / "t.parquet", "parquet")
    assert n == 3
    back = pq.read_table(tmp_path / "t.parquet")  # pyright: ignore[reportUnknownMemberType]
    assert back.column("reel_no").to_pylist() == ["012345", "054321", "000777"]


def test_export_subset_only_selected_rows(tmp_path: Path) -> None:
    n = exporters.export(_table(), tmp_path / "sel.parquet", "parquet", rows=[0, 2])
    assert n == 2
    back = pq.read_table(tmp_path / "sel.parquet")  # pyright: ignore[reportUnknownMemberType]
    assert back.column("name").to_pylist() == ["ACME", "GAMMA"]


def test_export_csv(tmp_path: Path) -> None:
    exporters.export(_table(), tmp_path / "t.csv", "csv")
    lines = (tmp_path / "t.csv").read_text(encoding="utf-8").splitlines()
    header = lines[0].replace('"', "")  # pyarrow may quote fields; compare quoting-agnostically
    assert header == "reel_no,name"
    assert "012345" in lines[1]  # leading zero preserved


def test_export_json_is_list_of_row_dicts(tmp_path: Path) -> None:
    exporters.export(_table(), tmp_path / "t.json", "json", rows=[1])
    data = json.loads((tmp_path / "t.json").read_text(encoding="utf-8"))
    assert data == [{"reel_no": "054321", "name": "BETA"}]


def test_export_feather_roundtrip(tmp_path: Path) -> None:
    exporters.export(_table(), tmp_path / "t.arrow", "feather")
    with pa.memory_map(str(tmp_path / "t.arrow"), "r") as source:
        back = pa.ipc.open_file(source).read_all()
    assert back.num_rows == 3
    assert back.column("reel_no").to_pylist() == ["012345", "054321", "000777"]


def test_export_xlsx_creates_sheet_with_header(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    n = exporters.export(_table(), tmp_path / "t.xlsx", "xlsx", sheet_name="props")
    assert n == 3
    wb = openpyxl.load_workbook(tmp_path / "t.xlsx")
    assert wb.sheetnames == ["props"]
    ws = wb["props"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header == ["reel_no", "name"]
    assert ws["A2"].value == "012345"  # stored as text


def test_export_rejects_unknown_format(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="unsupported format"):
        exporters.export(_table(), tmp_path / "t.zzz", "zzz")  # type: ignore[arg-type]


def _store(tmp_path: Path) -> TableStore:
    return parse_to_store(FIXTURE, tmp_path / "store")


def test_export_store_one_file_per_table(tmp_path: Path) -> None:
    store = _store(tmp_path)
    counts = exporters.export_store(store, tmp_path / "out", "csv")
    assert counts == {"assignments": 2, "assignors": 3, "assignees": 2, "properties": 4, "flat": 4}
    for name in ("assignments", "assignors", "assignees", "properties", "flat"):
        assert (tmp_path / "out" / f"{name}.csv").is_file()


def test_export_store_xlsx_is_single_multisheet_workbook(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    store = _store(tmp_path)
    counts = exporters.export_store(store, tmp_path / "out", "xlsx")
    book = tmp_path / "out" / "assignments.xlsx"
    assert book.is_file()
    wb = openpyxl.load_workbook(book)
    assert set(wb.sheetnames) == {"assignments", "assignors", "assignees", "properties", "flat"}
    assert counts["properties"] == 4


def test_write_workbook_multi_sheet(tmp_path: Path) -> None:
    sheets = {
        "alpha": pa.table({"x": ["1", "2"]}),
        "beta": pa.table({"y": ["a"]}),
    }
    counts = exporters.write_workbook(tmp_path / "wb.xlsx", sheets)
    assert counts == {"alpha": 2, "beta": 1}
    workbook = load_workbook(tmp_path / "wb.xlsx", read_only=True)
    try:
        assert workbook.sheetnames == ["alpha", "beta"]
    finally:
        workbook.close()
